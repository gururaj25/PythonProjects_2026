import os
import re
import openpyxl
from openpyxl.styles import Font

def extract_info_from_file(file_path):
    hostname = ''
    ip_address = ''
    oracle_versions = []
    oracle_present = 'No'

    try:
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
            content = f.read()

            hostname_match = re.search(r'Hostname:\s*(\S+)', content)
            ip_match = re.search(r'IP Address:\s*(\S+)', content)
            version_matches = re.findall(r'ORACLE_HOME\s+REG_SZ\s+.*?product\\([\d\.]+)\\', content)

            if hostname_match:
                hostname = hostname_match.group(1).strip()
            if ip_match:
                ip_address = ip_match.group(1).strip()
            if version_matches:
                oracle_versions = list(set(version_matches))
                oracle_present = 'Yes'
    except Exception as e:
        print(f"❌ Failed to process {file_path}: {e}")

    return {
        'Hostname': hostname,
        'IP Address': ip_address,
        'Oracle Present': oracle_present,
        'Oracle Versions': ':'.join(sorted(oracle_versions, reverse=True))
    }

def scan_directories_grouped(base_folder):
    grouped_data = {}

    for item in os.listdir(base_folder):
        folder_path = os.path.join(base_folder, item)
        if os.path.isdir(folder_path):
            folder_data = []
            for root, dirs, files in os.walk(folder_path):
                for file in files:
                    if file.lower() == "oracle_info.txt":
                        file_path = os.path.join(root, file)
                        print(f"Processing: {file_path}")
                        folder_data.append(extract_info_from_file(file_path))
            if folder_data:
                grouped_data[item] = folder_data
    return grouped_data

def save_grouped_data_to_excel(grouped_data, output_file):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet

    for folder_name, data_list in grouped_data.items():
        ws = wb.create_sheet(title=folder_name[:31])  # Max Excel sheet name length = 31
        headers = ['Hostname', 'IP Address', 'Oracle Present', 'Oracle Versions']
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)

        for data in data_list:
            ws.append([
                data['Hostname'],
                data['IP Address'],
                data['Oracle Present'],
                data['Oracle Versions']
            ])

    wb.save(output_file)
    print(f"✅ Excel report saved to: {os.path.abspath(output_file)}")

if __name__ == "__main__":
    BASE_DIR = r"D:\Projects_2025\Oracle_Projects_BidP"  # Update this path to your root folder
    OUTPUT_FILE = "Oracle_System_Report.xlsx"

    if not os.path.exists(BASE_DIR):
        print(f"❌ Base folder does not exist: {BASE_DIR}")
    else:
        grouped_info = scan_directories_grouped(BASE_DIR)
        if not grouped_info:
            print("⚠️ No Oracle_info.txt files found.")
        else:
            save_grouped_data_to_excel(grouped_info, OUTPUT_FILE)



