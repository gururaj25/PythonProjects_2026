import openpyxl
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from copy import copy
import os
from openpyxl.cell.cell import Cell

def extract_excel2_data_by_ip(excel2_path):
    data_by_ip = {}
    wb2 = load_workbook(excel2_path, data_only=True)

    for sheet in wb2.worksheets:
        headers = []
        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            if i == 0:
                headers = list(row)
                continue
            if not row or not row[1]:
                continue
            ip = str(row[1]).strip()
            if ip:
                data_by_ip[ip] = dict(zip(headers, row))

    return data_by_ip



def copy_sheet_with_formatting(src_sheet, target_sheet):
    for row in src_sheet.iter_rows():
        for cell in row:
            # Skip merged placeholder cells (MergedCell objects)
            if not isinstance(cell, Cell):
                continue
            # ✅ Use `cell.column` instead of `cell.col_idx`
            new_cell = target_sheet.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = copy(cell.number_format)
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)

    # ✅ Copy merged cell ranges
    for merged_range in src_sheet.merged_cells.ranges:
        target_sheet.merge_cells(str(merged_range))

    # ✅ Copy column widths
    for col in src_sheet.column_dimensions:
        target_sheet.column_dimensions[col].width = src_sheet.column_dimensions[col].width



def merge_excel1_with_excel2(excel1_path, excel2_path, output_path):
    wb1 = load_workbook(excel1_path)
    excel2_data = extract_excel2_data_by_ip(excel2_path)
    output_wb = Workbook()
    output_wb.remove(output_wb.active)  # Remove default sheet

    for sheet in wb1.worksheets:
        new_sheet = output_wb.create_sheet(title=sheet.title)
        copy_sheet_with_formatting(sheet, new_sheet)

        header_row = [cell.value for cell in sheet[1]]
        ip_col_index = header_row.index("IP Address") + 1

        # Find unique Excel2 keys
        extra_keys = set()
        for ip in excel2_data:
            extra_keys.update(excel2_data[ip].keys())
        extra_keys = sorted(k for k in extra_keys if k and k != "IP Address")

        # Add extra headers
        start_col = len(header_row) + 1
        for j, key in enumerate(extra_keys):
            cell = new_sheet.cell(row=1, column=start_col + j, value=f"E2_{key}")
            cell.font = copy(sheet.cell(row=1, column=1).font)  # Copy formatting

        # Go row by row
        for i, row in enumerate(sheet.iter_rows(min_row=2), start=2):
            ip = str(row[ip_col_index - 1].value).strip() if row[ip_col_index - 1].value else ""
            e2_data = excel2_data.get(ip, {})
            for j, key in enumerate(extra_keys):
                value = e2_data.get(key, "")
                new_sheet.cell(row=i, column=start_col + j, value=value)

    output_wb.save(output_path)
    print(f"✅ Final merged Excel saved with formatting: {os.path.abspath(output_path)}")

if __name__ == "__main__":
    # Set paths
    excel1_path = r"D:\Gururaj\Learning\Python\OracleResultScript\StationIPDetails_Master.xlsx"  # base file
    excel2_path = r"D:\Gururaj\Learning\Python\OracleResultScript\Oracle_System_Report.xlsx"   # additional info
    output_path = r"D:\Gururaj\Learning\Python\OracleResultScript\Merged_Oracle_With_All_Rows.xlsx"

    if not os.path.exists(excel1_path) or not os.path.exists(excel2_path):
        print("❌ Excel file path is incorrect. Please check.")
    else:
        merge_excel1_with_excel2(excel1_path, excel2_path, output_path)



